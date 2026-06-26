import flet as ft
import io
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from api.client import ApiClient, ApiError
from api.tickets import TicketsApi
from api.sessions import SessionsApi
from api.movies import MoviesApi
from api.halls import HallsApi
from state.app_state import AppState
from models.ticket import SalesStatistics


REPORTS = [
    ("paid_tickets", "Оплаченные билеты"),
    ("all_tickets", "Все билеты"),
    ("sessions_schedule", "Расписание сеансов"),
    ("movies_catalog", "Каталог фильмов"),
    ("revenue_by_hall", "Выручка по залам"),
]

PERIODS = [
    ("week", "Неделя", "7d"),
    ("month", "Месяц", "30d"),
    ("all", "Все время", "all"),
]


class AdminStatsView(ft.Column):
    def __init__(self, api_client: ApiClient, app_state: AppState):
        self._api_client = api_client
        self._app_state = app_state
        self._tickets_api = TicketsApi(api_client)
        self._sessions_api = SessionsApi(api_client)
        self._movies_api = MoviesApi(api_client)
        self._halls_api = HallsApi(api_client)
        self._stats: SalesStatistics | None = None

        self._progress = ft.ProgressBar(visible=False, bar_height=2)
        self._stats_container = ft.Container()

        def _make_period_items(report_key: str, fmt: str):
            items = []
            for period_key, period_label, _ in PERIODS:
                items.append(ft.MenuItemButton(
                    content=ft.Text(period_label),
                    on_click=lambda _, rk=report_key, pk=period_key, f=fmt: self._do_export(rk, pk, f),
                ))
            return items

        def _make_report_submenu(fmt: str):
            items = []
            for report_key, report_label in REPORTS:
                items.append(ft.SubmenuButton(
                    content=ft.Text(report_label),
                    controls=_make_period_items(report_key, fmt),
                ))
            return items

        self._menu_excel = ft.SubmenuButton(
            content=ft.Text("Excel"),
            controls=_make_report_submenu("excel"),
        )
        self._menu_txt = ft.SubmenuButton(
            content=ft.Text("TXT"),
            controls=_make_report_submenu("txt"),
        )
        self._menu_bar = ft.MenuBar(
            controls=[self._menu_excel, self._menu_txt],
        )

        super().__init__(
            scroll=ft.ScrollMode.AUTO,
            spacing=8,
            controls=[
                ft.Container(
                    padding=ft.padding.Padding(16, 16, 16, 0),
                    content=ft.Column(spacing=8, controls=[
                        ft.Text("Статистика продаж", size=24, weight=ft.FontWeight.BOLD),
                        ft.ResponsiveRow(
                            spacing=8,
                            controls=[
                                ft.Container(
                                    col={"xs": 12, "sm": 8},
                                    content=ft.Row(
                                        spacing=4,
                                        wrap=True,
                                        controls=[
                                            self._menu_bar,
                                            ft.Button(
                                                "Обновить",
                                                icon=ft.Icons.REFRESH,
                                                on_click=lambda _: self._load_stats(),
                                            ),
                                        ],
                                    ),
                                ),
                            ],
                        ),
                    ]),
                ),
                self._progress,
                ft.Container(
                    padding=16,
                    content=self._stats_container,
                ),
            ],
            expand=True,
        )

    def did_mount(self):
        self._load_stats()

    def _load_stats(self):
        self._progress.visible = True
        self.update()

        try:
            self._stats = self._tickets_api.get_sales_statistics()
            self._render()
        except ApiError as ex:
            self._show_snackbar(f"Ошибка: {ex.detail}")
        except Exception as ex:
            self._show_snackbar(f"Ошибка подключения: {ex}")
        finally:
            self._progress.visible = False
            self.update()

    def _render(self):
        s = self._stats
        if not s:
            return

        def stat_card(icon, label, value, color):
            return ft.Container(
                border_radius=12,
                bgcolor=ft.Colors.SURFACE_CONTAINER_HIGHEST,
                padding=20,
                content=ft.Column(
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    spacing=8,
                    controls=[
                        ft.Icon(icon, size=32, color=color),
                        ft.Text(str(value), size=28, weight=ft.FontWeight.BOLD),
                        ft.Text(label, size=13, color=ft.Colors.ON_SURFACE_VARIANT),
                    ],
                ),
            )

        self._stats_container.content = ft.Column(
            spacing=16,
            controls=[
                ft.ResponsiveRow(
                    spacing=12,
                    controls=[
                        ft.Container(col={"xs": 6, "sm": 4}, content=stat_card(ft.Icons.CONFIRMATION_NUMBER, "Всего билетов", s.total_tickets_sold, ft.Colors.PRIMARY)),
                        ft.Container(col={"xs": 6, "sm": 4}, content=stat_card(ft.Icons.PAID, "Оплачено", s.paid_tickets, ft.Colors.GREEN)),
                        ft.Container(col={"xs": 6, "sm": 4}, content=stat_card(ft.Icons.MONEY_OFF, "Не оплачено", s.unpaid_tickets, ft.Colors.ORANGE)),
                    ],
                ),
                ft.ResponsiveRow(
                    spacing=12,
                    controls=[
                        ft.Container(col={"xs": 6, "sm": 4}, content=stat_card(ft.Icons.THEATERS, "Всего сеансов", s.total_sessions, ft.Colors.BLUE)),
                        ft.Container(col={"xs": 6, "sm": 4}, content=stat_card(ft.Icons.PERCENT, "% оплаты", f"{s.payment_percentage}%", ft.Colors.TEAL)),
                        ft.Container(col={"xs": 6, "sm": 4}, content=stat_card(ft.Icons.ANALYTICS, "Среднее/сеанс", s.average_tickets_per_session, ft.Colors.PURPLE)),
                    ],
                ),
            ],
        )
        self.update()

    def _period_cutoff(self, period: str):
        if period == "week":
            return datetime.now() - timedelta(days=7)
        elif period == "month":
            return datetime.now() - timedelta(days=30)
        return None

    def _period_slug(self, period: str) -> str:
        for pk, pl, slug in PERIODS:
            if pk == period:
                return slug
        return "all"

    def _period_label(self, period: str) -> str:
        for pk, pl, _ in PERIODS:
            if pk == period:
                return pl
        return "Все время"

    def _load_sessions_map(self, period: str = "all"):
        cutoff = self._period_cutoff(period)
        sessions = []
        skip = 0
        limit = 100
        while True:
            batch = self._sessions_api.get_all(skip=skip, limit=limit)
            if not batch:
                break
            sessions.extend(batch)
            if len(batch) < limit:
                break
            skip += limit
        if cutoff:
            sessions = [s for s in sessions if s.datetime >= cutoff]
        return {s.id: s for s in sessions}

    def _load_halls_map(self):
        halls = self._halls_api.get_all()
        return {h.id: h.name for h in halls}

    def _load_movies_map_for_sessions(self, sessions):
        movie_ids = {s.movie_id for s in sessions}
        all_movies = self._movies_api.get_all(skip=0, limit=100)
        return {m.id: m for m in all_movies if m.id in movie_ids}

    def _collect_tickets(self, is_paid=None, period="all"):
        cutoff = self._period_cutoff(period)
        sessions_map = self._load_sessions_map(period)
        valid_session_ids = set(sessions_map.keys())

        tickets = []
        skip = 0
        limit = 100
        while True:
            kwargs = {"skip": skip, "limit": limit}
            if is_paid is not None:
                kwargs["is_paid"] = is_paid
            batch = self._tickets_api.get_all(**kwargs)
            if not batch:
                break
            tickets.extend(batch)
            if len(batch) < limit:
                break
            skip += limit

        tickets = [t for t in tickets if t.session_id in valid_session_ids]
        return tickets, sessions_map

    def _get_report_data(self, key: str, period: str):
        if key in ("paid_tickets", "all_tickets"):
            is_paid = True if key == "paid_tickets" else None
            tickets, sessions_map = self._collect_tickets(is_paid=is_paid, period=period)
            if not tickets:
                return None
            movies_map = self._load_movies_map_for_sessions(list(sessions_map.values()))
            halls_map = self._load_halls_map()
            return {"tickets": tickets, "sessions": sessions_map, "movies": movies_map, "halls": halls_map, "key": key}

        elif key == "sessions_schedule":
            sessions_map = self._load_sessions_map(period)
            sessions = list(sessions_map.values())
            if not sessions:
                return None
            movies_map = self._load_movies_map_for_sessions(sessions)
            halls_map = self._load_halls_map()
            return {"sessions": sessions, "movies": movies_map, "halls": halls_map, "key": key}

        elif key == "movies_catalog":
            movies = self._movies_api.get_all(skip=0, limit=100)
            return {"movies_list": movies, "key": key}

        elif key == "revenue_by_hall":
            tickets, sessions_map = self._collect_tickets(is_paid=True, period=period)
            if not tickets:
                return None
            halls_map = self._load_halls_map()
            return {"tickets": tickets, "sessions": sessions_map, "halls": halls_map, "key": key}

        return None

    def _do_export(self, report_key: str, period: str, fmt: str):
        if fmt == "excel":
            self._export_excel(report_key, period)
        else:
            self._export_txt(report_key, period)

    def _save_dir(self):
        from utils.file_utils import get_save_dir
        return get_save_dir("starmin")

    def _export_excel(self, report_key: str, period: str):
        self._progress.visible = True
        self.update()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            self._show_snackbar("Ошибка: openpyxl не установлен")
            self._progress.visible = False
            self.update()
            return

        try:
            data = self._get_report_data(report_key, period)
            if data is None:
                self._show_snackbar("Нет данных для экспорта")
                return

            key = data["key"]

            wb = Workbook()
            ws = wb.active

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            def style_header(ws_ref, headers):
                for col, h in enumerate(headers, 1):
                    c = ws_ref.cell(row=1, column=col, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = header_align
                    c.border = thin_border

            def style_cell(ws_ref, row, col, value, money=False):
                c = ws_ref.cell(row=row, column=col, value=value)
                c.border = thin_border
                if money:
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")

            if key in ("paid_tickets", "all_tickets"):
                ws.title = "Оплаченные билеты" if key == "paid_tickets" else "Все билеты"
                headers = ["ID билета", "Статус", "Дата сеанса", "Зал", "Фильм", "Возраст+", "Цена"]
                style_header(ws, headers)

                tickets = data["tickets"]
                sessions_map = data["sessions"]
                movies_map = data["movies"]
                halls_map = data["halls"]

                for row_idx, ticket in enumerate(tickets, 2):
                    session = sessions_map.get(ticket.session_id)
                    movie = movies_map.get(session.movie_id) if session else None
                    hall_name = halls_map.get(session.hall_id, "—") if session else "—"

                    dt = session.datetime.strftime("%d.%m.%Y %H:%M") if session else "—"
                    movie_title = movie.title if movie else "—"
                    age_r = f"{movie.age_restriction}+" if movie else "—"
                    status = "Оплачен" if ticket.is_paid else "Не оплачен"

                    style_cell(ws, row_idx, 1, ticket.id)
                    style_cell(ws, row_idx, 2, status)
                    style_cell(ws, row_idx, 3, dt)
                    style_cell(ws, row_idx, 4, hall_name)
                    style_cell(ws, row_idx, 5, movie_title)
                    style_cell(ws, row_idx, 6, age_r)
                    style_cell(ws, row_idx, 7, ticket.price, money=True)

                for col_letter, w in zip("ABCDEFG", [10, 14, 20, 15, 35, 12, 12]):
                    ws.column_dimensions[col_letter].width = w

            elif key == "sessions_schedule":
                ws.title = "Расписание сеансов"
                headers = ["ID сеанса", "Дата/Время", "Фильм", "Возраст+", "Зал", "Цена"]
                style_header(ws, headers)

                sessions = sorted(data["sessions"], key=lambda s: s.datetime)
                movies_map = data["movies"]
                halls_map = data["halls"]

                for row_idx, s in enumerate(sessions, 2):
                    movie = movies_map.get(s.movie_id)
                    hall_name = halls_map.get(s.hall_id, "—")

                    dt_str = s.datetime.strftime("%d.%m.%Y %H:%M")
                    movie_title = movie.title if movie else "—"
                    age_r = f"{movie.age_restriction}+" if movie else "—"

                    style_cell(ws, row_idx, 1, s.id)
                    style_cell(ws, row_idx, 2, dt_str)
                    style_cell(ws, row_idx, 3, movie_title)
                    style_cell(ws, row_idx, 4, age_r)
                    style_cell(ws, row_idx, 5, hall_name)
                    style_cell(ws, row_idx, 6, s.price, money=True)

                for col_letter, w in zip("ABCDEF", [10, 20, 35, 12, 15, 12]):
                    ws.column_dimensions[col_letter].width = w

            elif key == "movies_catalog":
                ws.title = "Каталог фильмов"
                headers = ["ID", "Название", "Жанр", "Длительность (мин)", "Возраст+"]
                style_header(ws, headers)

                for row_idx, m in enumerate(data["movies_list"], 2):
                    style_cell(ws, row_idx, 1, m.id)
                    style_cell(ws, row_idx, 2, m.title)
                    style_cell(ws, row_idx, 3, m.genre)
                    style_cell(ws, row_idx, 4, m.duration)
                    style_cell(ws, row_idx, 5, f"{m.age_restriction}+")

                for col_letter, w in zip("ABCDE", [8, 35, 18, 18, 12]):
                    ws.column_dimensions[col_letter].width = w

            elif key == "revenue_by_hall":
                ws.title = "Выручка по залам"
                headers = ["Зал", "Оплаченных билетов", "Сумма выручки"]
                style_header(ws, headers)

                tickets = data["tickets"]
                sessions_map = data["sessions"]
                halls_map = data["halls"]

                hall_stats: dict[int, dict] = defaultdict(lambda: {"count": 0, "revenue": 0.0})
                for t in tickets:
                    session = sessions_map.get(t.session_id)
                    if session:
                        hall_stats[session.hall_id]["count"] += 1
                        hall_stats[session.hall_id]["revenue"] += t.price

                for row_idx, (hid, stats) in enumerate(sorted(hall_stats.items()), 2):
                    hall_name = halls_map.get(hid, f"Зал {hid}")
                    style_cell(ws, row_idx, 1, hall_name)
                    style_cell(ws, row_idx, 2, stats["count"])
                    style_cell(ws, row_idx, 3, stats["revenue"], money=True)

                for col_letter, w in zip("ABC", [20, 22, 18]):
                    ws.column_dimensions[col_letter].width = w

            slug = self._period_slug(period)
            save_path = os.path.join(self._save_dir(), f"starmin_{key}_{slug}.xlsx")
            buf = io.BytesIO()
            wb.save(buf)
            wb.close()
            buf.seek(0)
            with open(save_path, "wb") as f:
                f.write(buf.getvalue())

            self._show_snackbar(f"Экспортировано: {save_path}")

        except ApiError as ex:
            self._show_snackbar(f"Ошибка API: {ex.detail}")
        except Exception as ex:
            self._show_snackbar(f"Ошибка: {ex}")
        finally:
            self._progress.visible = False
            self.update()

    def _export_txt(self, report_key: str, period: str):
        self._progress.visible = True
        self.update()

        try:
            data = self._get_report_data(report_key, period)
            if data is None:
                self._show_snackbar("Нет данных для экспорта")
                return

            key = data["key"]
            period_label = self._period_label(period)
            lines = []

            if key in ("paid_tickets", "all_tickets"):
                report_name = "Оплаченные билеты" if key == "paid_tickets" else "Все билеты"
                tickets = data["tickets"]
                sessions_map = data["sessions"]
                movies_map = data["movies"]
                halls_map = data["halls"]

                lines = [f"StarMIN Cinema — {report_name} ({period_label})", "=" * 60, ""]
                for ticket in tickets:
                    session = sessions_map.get(ticket.session_id)
                    movie = movies_map.get(session.movie_id) if session else None
                    hall_name = halls_map.get(session.hall_id, "—") if session else "—"

                    dt = session.datetime.strftime("%d.%m.%Y %H:%M") if session else "—"
                    title = movie.title if movie else "—"
                    age = f"{movie.age_restriction}+" if movie else "—"
                    status = "Оплачен" if ticket.is_paid else "Не оплачен"

                    lines.append(f"Билет #{ticket.id}")
                    lines.append(f"  Статус: {status}")
                    lines.append(f"  Дата/время: {dt}")
                    lines.append(f"  Зал: {hall_name}")
                    lines.append(f"  Фильм: {title}")
                    lines.append(f"  Возраст+: {age}")
                    lines.append(f"  Цена: {ticket.price} ₽")
                    lines.append("")

            elif key == "sessions_schedule":
                sessions = sorted(data["sessions"], key=lambda s: s.datetime)
                movies_map = data["movies"]
                halls_map = data["halls"]

                lines = [f"StarMIN Cinema — Расписание сеансов ({period_label})", "=" * 60, ""]
                for s in sessions:
                    movie = movies_map.get(s.movie_id)
                    hall_name = halls_map.get(s.hall_id, "—")
                    dt = s.datetime.strftime("%d.%m.%Y %H:%M")
                    title = movie.title if movie else "—"
                    age = f"{movie.age_restriction}+" if movie else "—"

                    lines.append(f"Сеанс #{s.id}")
                    lines.append(f"  Дата/время: {dt}")
                    lines.append(f"  Фильм: {title}")
                    lines.append(f"  Возраст+: {age}")
                    lines.append(f"  Зал: {hall_name}")
                    lines.append(f"  Цена: {s.price} ₽")
                    lines.append("")

            elif key == "movies_catalog":
                movies = data["movies_list"]
                lines = ["StarMIN Cinema — Каталог фильмов", "=" * 60, ""]
                for m in movies:
                    lines.append(f"Фильм #{m.id}")
                    lines.append(f"  Название: {m.title}")
                    lines.append(f"  Жанр: {m.genre}")
                    lines.append(f"  Длительность: {m.duration} мин")
                    lines.append(f"  Возраст+: {m.age_restriction}+")
                    lines.append("")

            elif key == "revenue_by_hall":
                tickets = data["tickets"]
                sessions_map = data["sessions"]
                halls_map = data["halls"]

                hall_stats: dict[int, dict] = defaultdict(lambda: {"count": 0, "revenue": 0.0})
                for t in tickets:
                    session = sessions_map.get(t.session_id)
                    if session:
                        hall_stats[session.hall_id]["count"] += 1
                        hall_stats[session.hall_id]["revenue"] += t.price

                lines = [f"StarMIN Cinema — Выручка по залам ({period_label})", "=" * 60, ""]
                for hid, stats in sorted(hall_stats.items()):
                    hall_name = halls_map.get(hid, f"Зал {hid}")
                    lines.append(f"Зал: {hall_name}")
                    lines.append(f"  Оплаченных билетов: {stats['count']}")
                    lines.append(f"  Сумма выручки: {stats['revenue']:.2f} ₽")
                    lines.append("")

            else:
                self._show_snackbar("Неизвестный тип отчёта")
                return

            slug = self._period_slug(period)
            save_path = os.path.join(self._save_dir(), f"starmin_{key}_{slug}.txt")
            with open(save_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

            self._show_snackbar(f"Экспортировано: {save_path}")

        except ApiError as ex:
            self._show_snackbar(f"Ошибка API: {ex.detail}")
        except Exception as ex:
            self._show_snackbar(f"Ошибка: {ex}")
        finally:
            self._progress.visible = False
            self.update()

    def _show_snackbar(self, msg: str):
        if self.page:
            sb = ft.SnackBar(ft.Text(msg))
            self.page.overlay.append(sb)
            sb.open = True
            self.page.update()
